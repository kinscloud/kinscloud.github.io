from openpyxl import load_workbook

workbook = load_workbook(filename=r'F:\kinscloud.github.io\Python\Excel\testxls.xlsx')
#print(workbook.sheetnames)
sheet = workbook.active
for row in sheet.rows:
#for row in sheet.iter_rows(min_row=2,max_row=3,min_col=1,max_col=3):
    for cell in row:
        print(cell.value)

#sheet['A1']='ids'
cell=sheet['A1']
cell.value='ids'
workbook.save(filename=r'F:\kinscloud.github.io\Python\Excel\testxls2.xlsx')